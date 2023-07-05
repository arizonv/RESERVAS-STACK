from rest_framework.response import Response
from rest_framework.views import APIView
from .serializers import UserSerializer,LoginSerializer,listSerializer,createUserSerializer
from rest_framework import status

from rest_framework import generics
from accounts.models import User
from django.contrib.auth import authenticate, login, logout

from rest_framework.authentication import SessionAuthentication,TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.authtoken.models import Token
import openpyxl


class LoginAPIView(APIView):

    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = authenticate(request, username=serializer.validated_data['username'], password=serializer.validated_data['password'])
        if user is not None:
            login(request, user)
            token, created = Token.objects.get_or_create(user=user)
            return Response({'id': user.id,
                             'usuario': user.username,
                             'admin': user.is_staff,
                             'token': token.key,
                             }, status=status.HTTP_200_OK)
        else:
            return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)


class UserLogout(APIView):
    authentication_classes = [TokenAuthentication, SessionAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        logout(request)
        return Response('Logout successfully')




class UserList(generics.ListAPIView):
    # authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    queryset = User.objects.all()
    serializer_class = listSerializer
    def get(self, request, *args, **kwargs):
        return self.list(request, *args, **kwargs)




class Register(generics.GenericAPIView):
    serializer_class = createUserSerializer

    def post(self, request, *args,  **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        return Response({
            "user": listSerializer(user).data,
            "message": "User Created Successfully.",
        })




# Importar los módulos necesarios
from openpyxl.styles import Font, Alignment, PatternFill

class UserReport(APIView):
    def get(self, request):
        # Obtener los usuarios de la base de datos
        users = User.objects.all()

        # Crear el archivo Excel
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Agregar encabezados a las columnas y aplicar estilos
        header_font = Font(name='Calibri', bold=True,size=12, color='FFFFFF')
        header_fill = PatternFill(fill_type='solid', start_color='3F3F3F', end_color='3F3F3F')
        header_alignment = Alignment(horizontal='center')

        worksheet['A1'] = 'ID'
        worksheet['A1'].font = header_font
        worksheet['A1'].fill = header_fill
        worksheet['A1'].alignment = header_alignment
        worksheet.column_dimensions['A'].width = 10
        
        worksheet['B1'] = 'Nombre de Usuario'
        worksheet['B1'].font = header_font
        worksheet['B1'].fill = header_fill
        worksheet['B1'].alignment = header_alignment
        worksheet.column_dimensions['B'].width = 22

        worksheet['C1'] = 'Nombre'
        worksheet['C1'].font = header_font
        worksheet['C1'].fill = header_fill
        worksheet['C1'].alignment = header_alignment
        worksheet.column_dimensions['C'].width = 22

        worksheet['D1'] = 'Correo electrónico'
        worksheet['D1'].font = header_font
        worksheet['D1'].fill = header_fill
        worksheet['D1'].alignment = header_alignment
        worksheet.column_dimensions['D'].width = 25

        worksheet['E1'] = 'Date Joined'
        worksheet['E1'].font = header_font
        worksheet['E1'].fill = header_fill
        worksheet['E1'].alignment = header_alignment
        worksheet.column_dimensions['E'].width = 25

        worksheet['F1'] = 'Admin'
        worksheet['F1'].font = header_font
        worksheet['F1'].fill = header_fill
        worksheet['F1'].alignment = header_alignment
        worksheet.column_dimensions['F'].width = 15

        # Agregar los datos de los usuarios a las filas
        row_num = 2
        data_alignment = Alignment(horizontal='left')
        for user in users:
            worksheet.cell(row=row_num, column=1, value=user.id).alignment = data_alignment
            worksheet.cell(row=row_num, column=2, value=user.username).alignment = data_alignment
            worksheet.cell(row=row_num, column=3, value=user.name).alignment = data_alignment
            worksheet.cell(row=row_num, column=4, value=user.email).alignment = data_alignment
            worksheet.cell(row=row_num, column=5, value=user.date_joined.strftime('%d/%m/%Y - %H:%M:%S')).alignment = data_alignment
            worksheet.cell(row=row_num, column=6, value=user.is_staff).alignment = data_alignment
            row_num += 1

        # Establecer el nombre del archivo y el tipo de contenido
        filename = 'reporte_usuarios.xlsx'
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename={}'.format(filename)

        # Guardar el archivo Excel en la respuesta HTTP
        workbook.save(response)

        return response
